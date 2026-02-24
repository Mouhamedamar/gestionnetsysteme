from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate, update_session_auth_hash
from django.contrib.auth.models import User
from django.db import connection
from .models import UserProfile
from .permissions import IsAdminOrTechnicienOrCommercial


@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    """
    Endpoint de connexion pour obtenir un token JWT
    """
    username = request.data.get('username')
    password = request.data.get('password')
    
    if not username or not password:
        return Response(
            {'error': 'Username et password requis'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    user = authenticate(username=username, password=password)
    
    if user is None:
        return Response(
            {'error': 'Identifiants invalides'},
            status=status.HTTP_401_UNAUTHORIZED
        )
    
    # Récupérer ou créer le profil utilisateur
    profile, created = UserProfile.objects.get_or_create(
        user=user,
        defaults={'role': 'commercial'}
    )
    
    # Déterminer le rôle
    role = profile.role if profile else 'commercial'
    
    # Pour les admins, vérifier is_staff
    if role == 'admin' and not user.is_staff:
        # Si le profil dit admin mais l'utilisateur n'est pas staff, mettre à jour
        user.is_staff = True
        user.save()
    
    refresh = RefreshToken.for_user(user)
    
    # Récupérer les permissions personnalisées
    page_permissions = profile.page_permissions if profile and profile.page_permissions else None
    
    return Response({
        'refresh': str(refresh),
        'access': str(refresh.access_token),
        'user': {
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'role': role,
            'is_staff': user.is_staff,
            'page_permissions': page_permissions,
            'profile': {
                'role': role,
                'phone': profile.phone if profile else None,
                'page_permissions': page_permissions
            }
        }
    }, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([AllowAny])
def logout(request):
    """
    Endpoint de déconnexion (blacklist du token)
    """
    try:
        refresh_token = request.data.get('refresh')
        if refresh_token:
            token = RefreshToken(refresh_token)
            token.blacklist()
            return Response(
                {'message': 'Déconnexion réussie'},
                status=status.HTTP_200_OK
            )
        return Response(
            {'error': 'Token de rafraîchissement requis'},
            status=status.HTTP_400_BAD_REQUEST
        )
    except Exception as e:
        return Response(
            {'error': 'Token invalide'},
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
def profile(request):
    """
    Endpoint pour récupérer et mettre à jour le profil utilisateur
    """
    user = request.user

    if request.method == 'GET':
        return Response({
            'id': user.id,
            'username': user.username,
            'email': user.email,
        }, status=status.HTTP_200_OK)

    elif request.method == 'PATCH':
        username = request.data.get('username')
        email = request.data.get('email')

        if not username or not email:
            return Response(
                {'error': 'Username et email sont requis'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Vérifier si le username est déjà pris
        if User.objects.filter(username=username).exclude(id=user.id).exists():
            return Response(
                {'error': 'Ce nom d\'utilisateur est déjà pris'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Vérifier si l'email est déjà pris
        if User.objects.filter(email=email).exclude(id=user.id).exists():
            return Response(
                {'error': 'Cet email est déjà utilisé'},
                status=status.HTTP_400_BAD_REQUEST
            )

        user.username = username
        user.email = email
        user.save()

        return Response({
            'id': user.id,
            'username': user.username,
            'email': user.email,
        }, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_password(request):
    """
    Endpoint pour changer le mot de passe
    """
    user = request.user
    current_password = request.data.get('current_password')
    new_password = request.data.get('new_password')
    confirm_password = request.data.get('confirm_password')

    if not current_password or not new_password or not confirm_password:
        return Response(
            {'error': 'Tous les champs sont requis'},
            status=status.HTTP_400_BAD_REQUEST
        )

    if not user.check_password(current_password):
        return Response(
            {'error': 'Mot de passe actuel incorrect'},
            status=status.HTTP_400_BAD_REQUEST
        )

    if new_password != confirm_password:
        return Response(
            {'error': 'Les nouveaux mots de passe ne correspondent pas'},
            status=status.HTTP_400_BAD_REQUEST
        )

    if len(new_password) < 8:
        return Response(
            {'error': 'Le nouveau mot de passe doit contenir au moins 8 caractères'},
            status=status.HTTP_400_BAD_REQUEST
        )

    user.set_password(new_password)
    user.save()
    update_session_auth_hash(request, user)

    return Response(
        {'message': 'Mot de passe changé avec succès'},
        status=status.HTTP_200_OK
    )


from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAdminUser, AllowAny
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from django.http import HttpResponse
from .models import Client, Prospect
from .serializers import ClientSerializer, ProspectSerializer, UserSerializer
import openpyxl
from io import BytesIO


@api_view(['GET'])
@permission_classes([AllowAny])
def check_installation(request):
    """
    Vérifie l'état de l'installation de l'application
    """
    checks = {
        'database': False,
        'migrations': False,
        'admin_exists': False,
        'server': True
    }
    
    messages = {
        'database': 'Base de données non accessible',
        'migrations': 'Migrations non vérifiées',
        'admin_exists': 'Aucun utilisateur admin trouvé',
        'server': 'Serveur accessible'
    }
    
    try:
        # Vérifier la base de données
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT 1")
            checks['database'] = True
            messages['database'] = 'Base de données accessible'
        except Exception as e:
            messages['database'] = f'Erreur de connexion: {str(e)}'
        
        # Vérifier si un admin existe
        try:
            admin_count = User.objects.filter(is_staff=True).count()
            checks['admin_exists'] = admin_count > 0
            messages['admin_exists'] = f'{admin_count} utilisateur(s) admin trouvé(s)' if admin_count > 0 else 'Aucun utilisateur admin trouvé'
        except Exception as e:
            messages['admin_exists'] = f'Erreur: {str(e)}'
        
        # Vérifier les migrations (simplifié - on vérifie juste si les tables existent)
        try:
            table_names = connection.introspection.table_names()
            if table_names:
                checks['migrations'] = True
                messages['migrations'] = f'{len(table_names)} table(s) trouvée(s)'
            else:
                messages['migrations'] = 'Aucune table trouvée - migrations nécessaires'
        except Exception as e:
            messages['migrations'] = f'Erreur: {str(e)}'
        
    except Exception as e:
        messages['database'] = f'Erreur générale: {str(e)}'
    
    return Response({
        'installed': True,  # Forcer à True pour désactiver le mode démo
        'checks': checks,
        'messages': messages
    }, status=status.HTTP_200_OK)


class ProspectViewSet(viewsets.ModelViewSet):
    """
    ViewSet pour la gestion des prospects (avant conversion en client).
    """
    queryset = Prospect.objects.all()
    serializer_class = ProspectSerializer
    permission_classes = [IsAdminOrTechnicienOrCommercial]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status']
    search_fields = ['name', 'email', 'phone', 'company']
    ordering_fields = ['created_at', 'name', 'status']
    ordering = ['-created_at']

    @action(detail=True, methods=['post'])
    def convert_to_client(self, request, pk=None):
        """
        Convertit un prospect en client.
        - Crée (ou réutilise) un client à partir des infos du prospect.
        - Marque le prospect comme converti.
        """
        prospect = self.get_object()
        if prospect.status == Prospect.STATUS_CONVERTED:
            return Response(
                {'detail': 'Ce prospect est déjà converti.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        client_defaults = {
            'name': prospect.name,
            'phone': prospect.phone,
            'email': prospect.email,
            'company': prospect.company,
            'client_type': Client.TYPE_CLIENT,
        }

        if prospect.email:
            client, created = Client.objects.get_or_create(
                email=prospect.email,
                defaults=client_defaults,
            )
        else:
            client = Client.objects.create(**client_defaults)
            created = True

        prospect.status = Prospect.STATUS_CONVERTED
        prospect.save(update_fields=['status'])

        client_data = ClientSerializer(client).data
        return Response(
            client_data,
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK
        )

    @action(detail=True, methods=['post'])
    def blacklist(self, request, pk=None):
        """
        Blacklister un prospect : crée ou réutilise un client is_blacklisted=True
        et marque le prospect comme perdu.
        """
        prospect = self.get_object()
        client_defaults = {
            'name': prospect.name,
            'phone': prospect.phone,
            'email': prospect.email,
            'company': prospect.company,
            'client_type': Client.TYPE_CLIENT,
            'is_blacklisted': True,
        }

        if prospect.email:
            client, _ = Client.objects.get_or_create(
                email=prospect.email,
                defaults=client_defaults,
            )
            # si le client existait déjà, on s'assure qu'il est bien blacklisté
            if not client.is_blacklisted:
                client.is_blacklisted = True
                client.save(update_fields=['is_blacklisted', 'updated_at'])
        else:
            client = Client.objects.create(**client_defaults)

        prospect.status = Prospect.STATUS_LOST
        prospect.save(update_fields=['status'])

        return Response(ClientSerializer(client).data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='import-excel', parser_classes=[MultiPartParser, FormParser])
    def import_excel(self, request):
        """
        Importer des prospects depuis un fichier Excel (.xlsx).
        Colonnes reconnues (noms possibles) :
        - nom / name
        - email / mail
        - téléphone / telephone / phone / tel
        - entreprise / company / societe
        - statut / status (new, contacted, converted, lost)
        """
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': 'Aucun fichier envoyé. Envoyez un fichier avec la clé \"file\".'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Format non supporté. Utilisez un fichier .xlsx.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            wb = openpyxl.load_workbook(filename=file, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            return Response(
                {'error': f'Fichier Excel invalide: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not rows:
            return Response(
                {'error': 'Le fichier est vide.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        header = [str(c).strip().lower() if c else '' for c in rows[0]]
        created = 0
        errors = []

        def col(row, name_aliases):
            for na in name_aliases:
                try:
                    idx = header.index(na)
                    if idx < len(row) and row[idx] is not None:
                        return str(row[idx]).strip() or None
                except ValueError:
                    continue
            return None

        for i, row in enumerate(rows[1:], start=2):
            if not any(v for v in row):
                continue

            name = col(row, ['nom', 'name', 'nom complet']) or ''

            email = col(row, ['email', 'mail'])
            phone = col(row, ['telephone', 'phone', 'téléphone', 'tel'])
            company = col(row, ['entreprise', 'company', 'societe'])
            status_str = (col(row, ['statut', 'status']) or '').lower()

            if status_str in ['new', 'nouveau']:
                status_value = Prospect.STATUS_NEW
            elif status_str in ['contacted', 'contacte', 'contacté']:
                status_value = Prospect.STATUS_CONTACTED
            elif status_str in ['converted', 'converti']:
                status_value = Prospect.STATUS_CONVERTED
            elif status_str in ['lost', 'perdu']:
                status_value = Prospect.STATUS_LOST
            else:
                status_value = Prospect.STATUS_NEW

            Prospect.objects.create(
                name=(name or 'Prospect importé')[:200],
                email=email or None,
                phone=phone or None,
                company=company or None,
                status=status_value,
            )
            created += 1

        wb.close()
        return Response(
            {
                'created': created,
                'message': f'{created} prospect(s) importé(s).',
                'errors': errors[:50],
            },
            status=status.HTTP_200_OK,
        )


@api_view(['POST'])
@permission_classes([AllowAny])
def setup_admin(request):
    """
    Crée le premier utilisateur administrateur (uniquement si aucun admin n'existe)
    """
    # Vérifier si un admin existe déjà
    if User.objects.filter(is_staff=True).exists():
        return Response(
            {'error': 'Un utilisateur administrateur existe déjà. Utilisez la page de connexion.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    username = request.data.get('username')
    email = request.data.get('email')
    password = request.data.get('password')
    
    if not username or not email or not password:
        return Response(
            {'error': 'Username, email et password sont requis'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    if len(password) < 8:
        return Response(
            {'error': 'Le mot de passe doit contenir au moins 8 caractères'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Vérifier si le username existe déjà
    if User.objects.filter(username=username).exists():
        return Response(
            {'error': 'Ce nom d\'utilisateur existe déjà'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Créer l'utilisateur admin
    try:
        admin = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            is_staff=True,
            is_superuser=True,
            is_active=True
        )
        
        return Response({
            'message': 'Utilisateur administrateur créé avec succès',
            'user': {
                'id': admin.id,
                'username': admin.username,
                'email': admin.email
            }
        }, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response(
            {'error': f'Erreur lors de la création: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


class ClientViewSet(viewsets.ModelViewSet):
    """
    ViewSet pour la gestion des clients et prospects.
    Filtres: client_type (PROSPECT, CLIENT), is_blacklisted.
    Actions: convert_to_client, blacklist, export_excel, import_excel.
    """
    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    permission_classes = [IsAdminOrTechnicienOrCommercial]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['client_type', 'is_blacklisted']
    search_fields = ['name', 'email', 'phone', 'company']
    ordering_fields = ['name', 'created_at', 'updated_at']
    ordering = ['-created_at']

    def get_queryset(self):
        """
        Par défaut : uniquement les vrais clients (exclut PROSPECT).
        Si ?all=1 ou ?include_prospects=1 : retourne tous les contacts du modèle Client.
        """
        qs = super().get_queryset()
        include_prospects = self.request.query_params.get('all') == '1' or \
                            self.request.query_params.get('include_prospects') == '1'
        if include_prospects:
            return qs
        return qs.exclude(client_type=Client.TYPE_PROSPECT)

    @action(detail=True, methods=['post'])
    def convert_to_client(self, request, pk=None):
        """Passe un prospect en client."""
        client = self.get_object()
        if client.client_type == Client.TYPE_CLIENT:
            return Response(
                {'detail': 'Ce contact est déjà un client.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        client.client_type = Client.TYPE_CLIENT
        client.is_blacklisted = False
        client.save(update_fields=['client_type', 'is_blacklisted', 'updated_at'])
        serializer = self.get_serializer(client)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def blacklist(self, request, pk=None):
        """Met le contact en blacklist (ne plus contacter)."""
        client = self.get_object()
        client.is_blacklisted = True
        client.save(update_fields=['is_blacklisted', 'updated_at'])
        serializer = self.get_serializer(client)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """Exporte la liste des clients/prospects en fichier Excel (filtres appliqués)."""
        queryset = self.filter_queryset(self.get_queryset())
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Clients'
        headers = [
            'ID', 'Nom', 'Prénom', 'Nom (famille)', 'Téléphone', 'Email', 'Adresse',
            'Entreprise', 'Type', 'Blacklisté', 'Observation',
            'RCCM', 'N° Immatriculation', 'NINEA', 'Créé le', 'Modifié le'
        ]
        ws.append(headers)
        for c in queryset:
            ws.append([
                c.id, c.name or '', c.first_name or '', c.last_name or '', c.phone or '', c.email or '',
                c.address or '', c.company or '', c.get_client_type_display() or c.client_type,
                'Oui' if c.is_blacklisted else 'Non', (c.observation or '')[:500],
                c.rccm_number or '', c.registration_number or '', c.ninea_number or '',
                c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else '',
                c.updated_at.strftime('%Y-%m-%d %H:%M') if c.updated_at else '',
            ])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="clients.xlsx"'
        return response

    @action(detail=False, methods=['post'], url_path='import-excel', parser_classes=[MultiPartParser, FormParser])
    def import_excel(self, request):
        """Importe des prospects/clients depuis un fichier Excel (.xlsx)."""
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': 'Aucun fichier envoyé. Envoyez un fichier avec la clé "file".'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Format non supporté. Utilisez un fichier .xlsx.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            wb = openpyxl.load_workbook(filename=file, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            return Response(
                {'error': f'Fichier Excel invalide: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not rows:
            return Response(
                {'error': 'Le fichier est vide.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        # Première ligne = en-têtes (on accepte des noms flexibles)
        header = [str(c).strip().lower() if c else '' for c in rows[0]]
        created = 0
        errors = []
        for i, row in enumerate(rows[1:], start=2):
            if not any(v for v in row):
                continue
            def col(name_aliases):
                for na in name_aliases:
                    try:
                        idx = header.index(na)
                        if idx < len(row) and row[idx] is not None:
                            return str(row[idx]).strip() or None
                    except ValueError:
                        continue
                return None
            name = col(['nom', 'name', 'nom complet', 'nom complet'])
            if not name:
                errors.append(f'Ligne {i}: nom manquant')
                continue
            phone = col(['telephone', 'phone', 'téléphone', 'tel'])
            email = col(['email', 'mail'])
            address = col(['adresse', 'address'])
            company = col(['entreprise', 'company', 'societe'])
            client_type = col(['type', 'client_type'])
            if client_type and client_type.upper() in ('CLIENT', 'CLIENTE'):
                client_type = Client.TYPE_CLIENT
            else:
                client_type = Client.TYPE_PROSPECT
            observation = col(['observation', 'observations', 'notes', 'note'])
            if email and Client.objects.filter(email=email).exists():
                continue  # évite doublon par email
            Client.objects.create(
                name=name[:200],
                phone=phone or None,
                email=email or None,
                address=address or None,
                company=company or None,
                client_type=client_type,
                observation=observation or None,
            )
            created += 1
        wb.close()
        return Response({
            'created': created,
            'message': f'{created} contact(s) importé(s).',
            'errors': errors[:50],
        })


class UserViewSet(viewsets.ModelViewSet):
    """
    ViewSet pour la gestion des utilisateurs
    """
    serializer_class = UserSerializer
    permission_classes = [IsAdminUser]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['username', 'email']
    ordering_fields = ['username', 'email', 'date_joined', 'is_staff']
    ordering = ['-date_joined']
    
    def get_queryset(self):
        """Optimiser les requêtes en préchargeant le profil"""
        try:
            return User.objects.select_related('profile').all().order_by('-date_joined')
        except Exception as e:
            # Si erreur avec select_related, utiliser le queryset de base
            return User.objects.all().order_by('-date_joined')

